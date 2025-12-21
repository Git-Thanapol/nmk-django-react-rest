import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime
from django.db.models import Sum, Q, F
from django.db.models.functions import Coalesce
from .models import Product, PurchaseItem, InvoiceItem # Ensure Product is imported

def get_thai_datetime():
    """Returns current datetime in Thai format: 2 ตุลาคม 2568 18:46 น."""
    now = datetime.now()
    thai_months = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    year = now.year + 543
    return f"{now.day} {thai_months[now.month]} {year} {now.strftime('%H:%M')} น."

def get_thai_month_year(date_obj):
    """Returns Month Year in Thai: ตุลาคม ปี พ.ศ. 2568"""
    if not date_obj: return ""
    thai_months = [
        "", "มกราคม", "กุมภาพันธ์", "มีนาคม", "เมษายน", "พฤษภาคม", "มิถุนายน",
        "กรกฎาคม", "สิงหาคม", "กันยายน", "ตุลาคม", "พฤศจิกายน", "ธันวาคม"
    ]
    return f"{thai_months[date_obj.month]} ปี พ.ศ. {date_obj.year + 543}"

def generate_purchase_tax_report(queryset, company, start_date, end_date, report_basis='create_date'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Tax Report"

    # ... (Keep Styles same as before) ...
    font_header = Font(name='Sarabun', size=14, bold=True)
    font_sub = Font(name='Sarabun', size=11)
    font_bold = Font(name='Sarabun', size=11, bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # --- 1. HEADER SECTION ---
    # ... (Keep rows 1 & 2 same as before) ...
    ws.merge_cells('A1:B1'); ws['A1'] = get_thai_datetime()
    ws.merge_cells('D1:F1'); ws['D1'] = str(company); ws['D1'].font = font_header; ws['D1'].alignment = align_center
    ws.merge_cells('D2:F2'); ws['D2'] = "รายงานภาษีซื้อ"; ws['D2'].font = font_bold; ws['D2'].alignment = align_center
    
    # NEW: Add subtitle to show filtering basis
    basis_text = "(ยื่นตามวันที่เอกสาร)" if report_basis == 'create_date' else "(ยื่นตามวันที่จ่ายภาษี)"
    ws['D2'].value = f"รายงานภาษีซื้อ {basis_text}"

    # Row 3
    ws.merge_cells('D3:F3')
    ws['D3'] = f"เดือนภาษี {get_thai_month_year(start_date)}" 
    ws['D3'].alignment = align_center

    tax_id = getattr(company, 'tax_id', '-')
    ws['H3'] = f"เลขประจำตัวผู้เสียภาษี {tax_id}"
    ws['H3'].alignment = align_right

    # --- 2. TABLE HEADERS (Keep same) ---
    headers = [
        ('A5:A6', 'ลำดับ', 5), ('B5:C5', 'ใบกำกับภาษี', 0), ('B6:B6', 'วัน/เดือน/ปี', 12), ('C6:C6', 'เลขที่', 15),
        ('D5:D6', 'ชื่อผู้ขายสินค้า/ผู้รับบริการ', 30), ('E5:E6', 'เลขประจำตัวผู้เสียภาษี\nของผู้ขายสินค้า', 20),
        ('F5:G5', 'สถานประกอบการ', 0), ('F6:F6', 'สนญ.', 8), ('G6:G6', 'สาขาที่', 8),
        ('H5:H6', 'มูลค่าสินค้าหรือบริการ', 15), ('I5:I6', 'จำนวนเงินภาษีมูลค่าเพิ่ม', 15),
    ]
    # ... (Keep Header generation loop same as before) ...
    for cell_range, text, width in headers:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(':')[0]]
        cell.value = text
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        for row in ws[cell_range]:
            for c in row: c.border = border_thin
        if width > 0: ws.column_dimensions[cell_range[0]].width = width

    ws.print_title_rows = '1:6'

    # --- 3. DATA POPULATION ---
    current_row = 7
    seq = 1
    total_value = 0
    total_vat = 0

    for po in queryset:
        # NOTE: Standard practice is to show the Document Date (order_date) 
        # even if we sorted by Tax Date.
        po_date = po.order_date 
        thai_date = f"{po_date.day:02d}/{po_date.month:02d}/{po_date.year+543}"
        
        vendor_name = po.vendor.name if po.vendor else "Unknown"
        vendor_tax = getattr(po.vendor, 'tax_id', '') 
        
        val = po.subtotal
        vat = po.tax_amount
        
        total_value += val
        total_vat += vat

        data = [
            (seq, 'center'),
            (thai_date, 'center'),
            (po.po_number, 'left'), 
            (vendor_name, 'left'),
            (vendor_tax, 'center'),
            ('X', 'center'), ('', 'center'),
            (val, 'number'), (vat, 'number'),
        ]
        
        # ... (Keep Data writing loop same as before) ...
        col_indices = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for i, (value, style_type) in enumerate(data):
            cell = ws[f"{col_indices[i]}{current_row}"]
            cell.value = value
            cell.font = font_sub
            cell.border = border_thin
            if style_type == 'number':
                cell.number_format = '#,##0.00'
                cell.alignment = align_right
            elif style_type == 'center':
                cell.alignment = align_center
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        current_row += 1
        seq += 1

    # --- 4. TOTALS ROW (Keep same) ---
    # ... (Use your existing total row code) ...
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'] = "รวมทั้งสิ้น"
    ws[f'A{current_row}'].alignment = align_right
    ws[f'A{current_row}'].font = font_bold
    ws[f'A{current_row}'].border = border_thin
    for col in ['B','C','D','E','F','G']: ws[f'{col}{current_row}'].border = border_thin

    for col, val in [('H', total_value), ('I', total_vat)]:
        c = ws[f'{col}{current_row}']
        c.value = val
        c.font = font_bold
        c.number_format = '#,##0.00'
        c.alignment = align_right
        c.border = border_thin

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Purchase_Tax_{report_basis}_{start_date}.xlsx"'
    wb.save(response)
    return response

def generate_sales_tax_report(queryset, company, start_date, end_date, report_basis='create_date'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Tax Report"

    # ... (Styles) ...
    font_header = Font(name='Sarabun', size=14, bold=True)
    font_sub = Font(name='Sarabun', size=11)
    font_bold = Font(name='Sarabun', size=11, bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # ... (Headers) ...
    ws.merge_cells('A1:B1'); ws['A1'] = get_thai_datetime()
    ws.merge_cells('D1:F1'); ws['D1'] = str(company); ws['D1'].font = font_header; ws['D1'].alignment = align_center
    
    ws.merge_cells('D2:F2')
    basis_text = "(ยื่นตามวันที่เอกสาร)" if report_basis == 'create_date' else "(ยื่นตามวันที่จ่ายภาษี)"
    ws['D2'].value = f"รายงานภาษีขาย {basis_text}"
    ws['D2'].font = font_bold; ws['D2'].alignment = align_center

    ws.merge_cells('D3:F3')
    ws['D3'] = f"เดือนภาษี {get_thai_month_year(start_date)}" 
    ws['D3'].alignment = align_center

    tax_id = getattr(company, 'tax_id', '-')
    ws['H3'] = f"เลขประจำตัวผู้เสียภาษี {tax_id}"
    ws['H3'].alignment = align_right

    # ... (Table Headers - Same as before) ...
    headers = [
        ('A5:A6', 'ลำดับ', 5), ('B5:C5', 'ใบกำกับภาษี', 0), ('B6:B6', 'วัน/เดือน/ปี', 12), ('C6:C6', 'เลขที่', 15),
        ('D5:D6', 'ชื่อผู้ซื้อสินค้า/ผู้รับบริการ', 30), ('E5:E6', 'เลขประจำตัวผู้เสียภาษี\nของผู้ซื้อสินค้า', 20),
        ('F5:G5', 'สถานประกอบการ', 0), ('F6:F6', 'สนญ.', 8), ('G6:G6', 'สาขาที่', 8),
        ('H5:H6', 'มูลค่าสินค้าหรือบริการ', 15), ('I5:I6', 'จำนวนเงินภาษีมูลค่าเพิ่ม', 15),
    ]
    for cell_range, text, width in headers:
        ws.merge_cells(cell_range)
        cell = ws[cell_range.split(':')[0]]
        cell.value = text
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        for row in ws[cell_range]:
            for c in row: c.border = border_thin
        if width > 0: ws.column_dimensions[cell_range[0]].width = width

    ws.print_title_rows = '1:6'

    # ... (Data) ...
    current_row = 7
    seq = 1
    total_value = 0
    total_vat = 0

    for inv in queryset:
        inv_date = inv.invoice_date
        thai_date = f"{inv_date.day:02d}/{inv_date.month:02d}/{inv_date.year+543}"
        
        cust_name = inv.recipient_name if inv.recipient_name else (inv.vendor.name if inv.vendor else "เงินสด/ไม่ระบุ")
        cust_tax = inv.vendor.tax_id if (inv.vendor and hasattr(inv.vendor, 'tax_id')) else ""
        
        val = inv.subtotal
        vat = inv.tax_amount
        total_value += val
        total_vat += vat

        data = [
            (seq, 'center'),
            (thai_date, 'center'),
            (inv.invoice_number, 'left'),
            (cust_name, 'left'),
            (cust_tax, 'center'),
            ('X', 'center'), ('', 'center'),
            (val, 'number'), (vat, 'number'),
        ]

        # ... (Write Data Loop) ...
        col_indices = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for i, (value, style_type) in enumerate(data):
            cell = ws[f"{col_indices[i]}{current_row}"]
            cell.value = value
            cell.font = font_sub
            cell.border = border_thin
            if style_type == 'number':
                cell.number_format = '#,##0.00'
                cell.alignment = align_right
            elif style_type == 'center':
                cell.alignment = align_center
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        current_row += 1
        seq += 1

    # ... (Totals) ...
    ws.merge_cells(f'A{current_row}:G{current_row}')
    ws[f'A{current_row}'] = "รวมทั้งสิ้น"
    ws[f'A{current_row}'].alignment = align_right
    ws[f'A{current_row}'].font = font_bold
    ws[f'A{current_row}'].border = border_thin
    for col in ['B','C','D','E','F','G']: ws[f'{col}{current_row}'].border = border_thin

    for col, val in [('H', total_value), ('I', total_vat)]:
        c = ws[f'{col}{current_row}']
        c.value = val
        c.font = font_bold
        c.number_format = '#,##0.00'
        c.alignment = align_right
        c.border = border_thin

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Sales_Tax_{report_basis}_{start_date}.xlsx"'
    wb.save(response)
    return response

def generate_stock_report(company, start_date, end_date):
    """
    Generates Stock Report showing movement within range and absolute current balance.
    Formula: Actual Stock = All Time Buy - All Time Sell (Allows negative results)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    # --- Styles (Reusing your standard styles) ---
    font_header = Font(name='Sarabun', size=14, bold=True)
    font_sub = Font(name='Sarabun', size=11)
    font_bold = Font(name='Sarabun', size=11, bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    # --- 1. HEADER SECTION ---
    # Row 1
    ws.merge_cells('A1:B1')
    ws['A1'] = get_thai_datetime()
    ws['A1'].alignment = align_left
    
    ws.merge_cells('C1:E1')
    ws['C1'] = str(company)
    ws['C1'].font = font_header
    ws['C1'].alignment = align_center

    ws['F1'] = "หน้า 1"
    ws['F1'].alignment = align_right

    # Row 2
    ws.merge_cells('C2:E2')
    ws['C2'] = "รายงานสินค้าคงเหลือ" # Stock Report
    ws['C2'].font = font_bold
    ws['C2'].alignment = align_center

    ws['F2'] = "สำนักงานใหญ่"
    ws['F2'].alignment = align_right

    # Row 3
    ws.merge_cells('C3:E3')
    # Custom Date Range Text
    start_thai = f"{start_date.day}/{start_date.month}/{start_date.year+543}"
    end_thai = f"{end_date.day}/{end_date.month}/{end_date.year+543}"
    ws['C3'] = f"ข้อมูลตั้งแต่วันที่ {start_thai} ถึง {end_thai}" 
    ws['C3'].alignment = align_center

    tax_id = getattr(company, 'tax_id', '-')
    ws['F3'] = f"เลขประจำตัวผู้เสียภาษี {tax_id}"
    ws['F3'].alignment = align_right

    # --- 2. TABLE HEADERS ---
    headers = [
        ('A5:A6', 'ลำดับ', 8),
        ('B5:B6', 'รหัสสินค้า (SKU)', 20),
        ('C5:C6', 'ชื่อสินค้า', 40),
        ('D5:D5', 'ความเคลื่อนไหว (ช่วงเวลา)', 0), # Parent Header
        ('D6:D6', 'รับเข้า', 15),
        ('E6:E6', 'ขายออก', 15),
        ('F5:F6', 'คงเหลือปัจจุบัน\n(ทั้งหมด)', 20),
    ]

    # Merge Parent Header "Movement"
    ws.merge_cells('D5:E5')
    ws['D5'].alignment = align_center
    ws['D5'].font = font_bold
    ws['D5'].border = border_thin

    for cell_range, text, width in headers:
        # Skip if merged manually above
        if cell_range != 'D5:D5':
            ws.merge_cells(cell_range)
        
        top_left = cell_range.split(':')[0]
        cell = ws[top_left]
        cell.value = text
        cell.font = font_bold
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Apply borders
        for row in ws[cell_range]:
            for c in row:
                c.border = border_thin
        
        # Set Width
        if width > 0:
            col_letter = top_left[0]
            ws.column_dimensions[col_letter].width = width

    ws.print_title_rows = '1:6'

    # --- 3. DATA CALCULATION & POPULATION ---
    current_row = 7
    seq = 1

    # Fetch all active products
    products = Product.objects.filter(is_active=True).order_by('name')

    for product in products:
        # A. Movement within Date Range
        # Note: Filter by company to ensure multi-company isolation
        range_receive = PurchaseItem.objects.filter(
            product=product,
            #purchase_order__company=company,
            purchase_order__order_date__range=[start_date, end_date]
        ).aggregate(sum_qty=Coalesce(Sum('quantity'), 0))['sum_qty']

        range_sales = InvoiceItem.objects.filter(
            product=product,
            #invoice__company=company,
            invoice__status='BILLED', # Only count Billed sales
            invoice__invoice_date__range=[start_date, end_date]
        ).aggregate(sum_qty=Coalesce(Sum('quantity'), 0))['sum_qty']

        # B. Actual Stock (All Time)
        # Formula: Total In - Total Out
        all_time_in = PurchaseItem.objects.filter(
            product=product,
            #purchase_order__company=company
        ).aggregate(sum_qty=Coalesce(Sum('quantity'), 0))['sum_qty']

        all_time_out = InvoiceItem.objects.filter(
            product=product,
            #invoice__company=company,
            #invoice__status='BILLED'
        ).aggregate(sum_qty=Coalesce(Sum('quantity'), 0))['sum_qty']

        actual_stock = all_time_in - all_time_out

        # Skip rows if no movement AND no stock (optional, keeps report clean)
        if range_receive == 0 and range_sales == 0 and actual_stock == 0:
            continue

        # Write Row
        row_data = [
            (seq, 'center'),
            (product.sku if hasattr(product, 'sku') else '-', 'left'), # Adjust field name if needed
            (product.name, 'left'),
            (range_receive, 'number'),
            (range_sales, 'number'),
            (actual_stock, 'number_bold'), # Bold for emphasis
        ]

        col_indices = ['A', 'B', 'C', 'D', 'E', 'F']

        for i, (val, style) in enumerate(row_data):
            cell = ws[f"{col_indices[i]}{current_row}"]
            cell.value = val
            cell.border = border_thin
            cell.font = font_sub

            if 'number' in style:
                cell.number_format = '#,##0'
                cell.alignment = align_right
                if style == 'number_bold':
                    cell.font = font_bold
                    # Highlight negative stock in Red
                    if val < 0:
                        cell.font = Font(name='Sarabun', size=11, bold=True, color="FF0000")
            elif style == 'center':
                cell.alignment = align_center
            else:
                cell.alignment = align_left

        current_row += 1
        seq += 1

    # --- Response ---
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Stock_Report_{start_date}.xlsx"'
    wb.save(response)
    return response



def generate_combined_tax_report(po_qs, inv_qs, company_name, start_date, end_date, report_basis='create_date'):
    """
    Combines Purchase and Sales into one chronological report.
    Sorting Priority: 
    1. Date (Create or Tax depending on mode)
    2. Doc Type (Purchase comes before Sales)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined VAT Report"

    # --- Styles ---
    font_header = Font(name='Sarabun', size=14, bold=True)
    font_sub = Font(name='Sarabun', size=11)
    font_bold = Font(name='Sarabun', size=11, bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # --- 1. HEADER ---
    ws.merge_cells('A1:B1'); ws['A1'] = get_thai_datetime()
    ws.merge_cells('E1:G1'); ws['E1'] = str(company_name); ws['E1'].font = font_header; ws['E1'].alignment = align_center
    ws.merge_cells('E2:G2')
    basis_text = "(ยื่นตามวันที่เอกสาร)" if report_basis == 'create_date' else "(ยื่นตามวันที่จ่ายภาษี)"
    ws['E2'] = f"รายงานสรุปภาษีซื้อ - ขาย {basis_text}"
    ws['E2'].font = font_bold; ws['E2'].alignment = align_center
    
    ws.merge_cells('E3:G3')
    ws['E3'] = f"เดือนภาษี {get_thai_month_year(start_date)}"
    ws['E3'].alignment = align_center

    # --- 2. TABLE HEADERS ---
    # Structure: Date | Doc No | Type | Name | Tax ID | Buy Base | Buy VAT | Sell Base | Sell VAT
    headers = [
        ('A5:A6', 'ลำดับ', 5),
        ('B5:B6', 'วัน/เดือน/ปี', 12),
        ('C5:C6', 'เลขที่เอกสาร', 15),
        ('D5:D6', 'ประเภท', 10),
        ('E5:E6', 'ชื่อผู้ซื้อ / ผู้ขาย', 30),
        ('F5:F6', 'เลขผู้เสียภาษี', 18),
        ('G5:H5', 'ภาษีซื้อ (Purchase Tax)', 0),
        ('G6:G6', 'มูลค่า', 15),
        ('H6:H6', 'ภาษี', 12),
        ('I5:J5', 'ภาษีขาย (Sales Tax)', 0),
        ('I6:I6', 'มูลค่า', 15),
        ('J6:J6', 'ภาษี', 12),
    ]

    for cell_range, text, width in headers:
        if ':' in cell_range and cell_range.split(':')[0] != cell_range.split(':')[1]:
            ws.merge_cells(cell_range)
        
        cell = ws[cell_range.split(':')[0]]
        cell.value = text
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = border_thin
        
        # Apply borders to all cells in range
        for row in ws[cell_range]:
            for c in row: c.border = border_thin
            
        if width > 0: ws.column_dimensions[cell_range[0]].width = width

    ws.print_title_rows = '1:6'

    # --- 3. DATA PREPARATION & SORTING ---
    combined_data = []

    # Process Purchases (Priority 1)
    for po in po_qs:
        sort_date = po.order_date if report_basis == 'create_date' else po.tax_sender_date
        
        combined_data.append({
            'sort_date': sort_date,
            'priority': 1, # Purchase comes first
            'doc_no': po.po_number,
            'type': 'ซื้อ (Buy)',
            'name': po.vendor.name if po.vendor else '-',
            'tax_id': getattr(po.vendor, 'tax_id', '-'),
            'buy_base': po.subtotal,
            'buy_vat': po.tax_amount,
            'sell_base': 0,
            'sell_vat': 0
        })

    # Process Invoices (Priority 2)
    for inv in inv_qs:
        sort_date = inv.invoice_date if report_basis == 'create_date' else inv.tax_sender_date
        
        # Determine name logic
        name = inv.recipient_name if inv.recipient_name else (inv.vendor.name if inv.vendor else "เงินสด")
        # Determine Tax ID logic
        tax_id = inv.vendor.tax_id if (inv.vendor and hasattr(inv.vendor, 'tax_id')) else ""

        combined_data.append({
            'sort_date': sort_date,
            'priority': 2, # Invoice comes second
            'doc_no': inv.invoice_number,
            'type': 'ขาย (Sell)',
            'name': name,
            'tax_id': tax_id,
            'buy_base': 0,
            'buy_vat': 0,
            'sell_base': inv.subtotal,
            'sell_vat': inv.tax_amount
        })

    # SORTING: Date ASC -> Priority ASC -> Doc No ASC
    combined_data.sort(key=lambda x: (x['sort_date'], x['priority'], x['doc_no']))

    # --- 4. POPULATE EXCEL ---
    current_row = 7
    seq = 1
    total_buy_base = 0
    total_buy_vat = 0
    total_sell_base = 0
    total_sell_vat = 0

    for item in combined_data:
        # Format Date
        d = item['sort_date']
        thai_date = f"{d.day:02d}/{d.month:02d}/{d.year+543}" if d else "-"

        # Accumulate Totals
        total_buy_base += item['buy_base']
        total_buy_vat += item['buy_vat']
        total_sell_base += item['sell_base']
        total_sell_vat += item['sell_vat']

        row_cells = [
            (seq, 'center'),
            (thai_date, 'center'),
            (item['doc_no'], 'left'),
            (item['type'], 'center'),
            (item['name'], 'left'),
            (item['tax_id'], 'center'),
            (item['buy_base'], 'number'),
            (item['buy_vat'], 'number'),
            (item['sell_base'], 'number'),
            (item['sell_vat'], 'number'),
        ]

        col_indices = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
        for i, (val, style) in enumerate(row_cells):
            cell = ws[f"{col_indices[i]}{current_row}"]
            cell.value = val
            cell.border = border_thin
            cell.font = font_sub

            if style == 'number':
                cell.number_format = '#,##0.00'
                if val == 0: cell.font = Font(color="D3D3D3") # Gray out zeros
                cell.alignment = align_right
            elif style == 'center':
                cell.alignment = align_center
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')

        current_row += 1
        seq += 1

    # --- 5. TOTALS ROW ---
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "รวมทั้งสิ้น"
    ws[f'A{current_row}'].alignment = align_right
    ws[f'A{current_row}'].font = font_bold
    
    totals = [total_buy_base, total_buy_vat, total_sell_base, total_sell_vat]
    cols = ['G', 'H', 'I', 'J']
    
    # Border for "Total" Label
    for c in ['A','B','C','D','E','F']: ws[f'{c}{current_row}'].border = border_thin

    # Fill Total Values
    for i, val in enumerate(totals):
        cell = ws[f'{cols[i]}{current_row}']
        cell.value = val
        cell.number_format = '#,##0.00'
        cell.font = font_bold
        cell.alignment = align_right
        cell.border = border_thin

    # --- 6. NET VAT SUMMARY (Optional but helpful) ---
    current_row += 2
    ws[f'H{current_row}'] = "ภาษีขายสุทธิ:"
    ws[f'I{current_row}'] = total_sell_vat
    
    ws[f'H{current_row+1}'] = "หัก ภาษีซื้อ:"
    ws[f'I{current_row+1}'] = total_buy_vat
    
    ws[f'H{current_row+2}'] = "ภาษีที่ต้องชำระ:"
    net_vat = total_sell_vat - total_buy_vat
    cell = ws[f'I{current_row+2}']
    cell.value = net_vat
    cell.number_format = '#,##0.00'
    cell.font = font_bold
    if net_vat < 0: cell.font = Font(bold=True, color="FF0000") # Red if refund

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Combined_VAT_{report_basis}_{start_date}.xlsx"'
    wb.save(response)
    return response