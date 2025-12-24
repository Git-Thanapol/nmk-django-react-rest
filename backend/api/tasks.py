import threading
from django.core.files.base import ContentFile
import openpyxl
from io import BytesIO
from datetime import datetime
from openpyxl.styles import Font, PatternFill

# Import your existing logic
from .utils_processors import process_tiktok_orders, process_shopee_orders, process_lazada_orders
from .legacy_files.utils_import_core import universal_invoice_import
from .models import ImportLog, Company

def run_import_background(log_id, file_path, company_id, user_id, platform):
    try:
        # 1. Get Log Entry
        log = ImportLog.objects.get(id=log_id)
        log.status = 'PROCESSING'
        log.save()

        # 2. Process File (Reuse your existing logic)
        header_df = None
        items_df = None
        target_platform_name = ''

        if platform == 'tiktok':
            header_df, items_df = process_tiktok_orders(file_path)
            target_platform_name = 'TikTok Shop'
        elif platform == 'shopee':
            header_df, items_df = process_shopee_orders(file_path)
            target_platform_name = 'Shopee'
        elif platform == 'lazada':
            header_df, items_df = process_lazada_orders(file_path)
            target_platform_name = 'Lazada'

        # 3. Run Universal Import
        result = universal_invoice_import(
            header_df, items_df, company_id, user_id, platform_name=target_platform_name
        )

        # 4. Update Stats
        log.total_records = result.get('imported', 0) + result.get('failed', 0)
        log.success_count = result.get('imported', 0)
        log.failed_count = result.get('failed', 0)

        # 5. Generate Error File if needed
        if result['failed'] > 0:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Import Errors"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            
            headers = ["Excel Row", "Order ID", "Error Message", "Suggestion"]
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill

            for err in result['error_log']:
                if isinstance(err, dict):
                    row_idx = err.get('row_index', '-')
                    order_id = err.get('order_id', '-')
                    reason = err.get('reason', '-')
                else:
                    row_idx, order_id, reason = "-", "Unknown", str(err)
                
                suggestion = ""
                if "numeric field overflow" in reason: suggestion = "Amount too large."
                elif "unique constraint" in reason: suggestion = "Order ID exists."
                
                ws.append([row_idx, order_id, reason, suggestion])

            # Save to Memory buffer
            virtual_workbook = BytesIO()
            wb.save(virtual_workbook)
            
            # Save to Model
            filename = f"error_report_{log_id}.xlsx"
            log.error_file.save(filename, ContentFile(virtual_workbook.getvalue()))
            log.status = 'COMPLETED_WITH_ERRORS'
        else:
            log.status = 'COMPLETED'

    except Exception as e:
        log.status = 'FAILED'
        # You might want to save the exception message to a note field
        print(f"Background Task Failed: {e}")

    finally:
        log.save()
        # Clean up temp file
        import os
        if os.path.exists(file_path):
            os.remove(file_path)